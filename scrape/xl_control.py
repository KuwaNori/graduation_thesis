#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
単語リスト作成のためのExcelのコントロール
"""

import openpyxl
import pprint
from collections import Counter

def test(xlfile):
    xlfile = 'excel/category5.xlsx'
    wb = openpyxl.load_workbook(xlfile)
    wsSheet = wb['Sheet1']
    print(wsSheet.max_row)

def getURLs(xlfile):
    wb = openpyxl.load_workbook(xlfile)
    wsURLs = wb['URLs']
    max_row = wsURLs.max_row
    urls = []
    for i in range(1, max_row+1):
        urls.append(wsURLs.cell(i,1).value)
    return urls

def addData(words,xlfile):
    wb = openpyxl.load_workbook(xlfile)
    wsIndex = wb['index']
    cnt_list = Counter(words)
    keys_list = cnt_list.keys()
    max_row = wsIndex.max_row
    # 既にxlにある値の出現回数の追加
    for i in range(1, max_row+1):
        word_incell = wsIndex.cell(i,1).value
        count = wsIndex.cell(i,2).value
        if word_incell in keys_list:
            if count != 0:
                wsIndex.cell(row=i, column=2, value=count + cnt_list[word_incell])
            cnt_list.pop(word_incell)
    if len(cnt_list) < 1:
        wb.save(xlfile)
        return False
    # xlに入っていなかった値の追加
    for i, key in zip(range(max_row+1, max_row+1+len(cnt_list)), cnt_list):
        wsIndex.cell(row=i,column=1, value = key)
        wsIndex.cell(row=i,column=2, value = cnt_list[key])
    wb.save(xlfile)

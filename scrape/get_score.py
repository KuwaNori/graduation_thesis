#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
ニュース記事の分類、スコア出しを行う
"""

import openpyxl
import scrape_morph as morph
from collections import Counter

def test_score():
    test_val = input("URL:")
    words = morph.getNouns(test_val)
    words = Counter(words)
    if len(words) == 0:
        print("We cannot find words")
        exit()
    print(words)
    categories_names =['貧困','飢餓','健康','教育','ジェンダー','環境']
    xlfiles = ['category1.xlsx', 'category2.xlsx', 'category3.xlsx', 'category4.xlsx', 'category5.xlsx', 'category13.xlsx',]
    feild = []
    for xl in xlfiles:
        print(xl)
        wb = openpyxl.load_workbook('excel/'+xl)
        wsIndex = wb['index']
        pre_words = {}
        for i in range(1,100):
            pre_words[wsIndex.cell(i,1).value] = wsIndex.cell(i,2).value
        feild.append(pre_words)
        print("done")
    sum_all = 0
    cate_score = []
    if 0 in cate_score:
        print("score is 0")
        exit()
    for i in range(len(xlfiles)):
        amount = 0
        for word in words:
            if word in feild[i]:
                amount +=feild[i][word] * words[word]
        cate_score.append(amount)
    print(cate_score)
    sum_all = sum(cate_score)
    for i,s in zip(categories_names,cate_score):
        print('{} : {:.1%}'.format( i, s/sum_all))

test_score()

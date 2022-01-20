#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import pprint
from collections import Counter

def new_value(t,last_v):
    values = [2330,3072,2175,2241,2385,3113,3002,2705,2326,2181,2443,2455,3208,3573,3242,2326]
    b = values[t-2]
    new = round((3000*last_v)/b)
    return new

xlfile = 'excel/final_words.xlsx'
wb = openpyxl.load_workbook(xlfile)
wsSheet = wb['pattern6']
max_row = wsSheet.max_row
for t in range(17,18):
    for i in range(1,max_row+1):
        value = wsSheet.cell(i,t).value
        if value is None:
            continue
        new = new_value(t,value)
        print(str(value)+" : "+str(new))
        wsSheet.cell(row=i,column=t,value=new)
wb.save(xlfile)
print("complete")

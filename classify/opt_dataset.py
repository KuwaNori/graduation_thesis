#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
import pprint
from collections import Counter

xlfile = 'excel/final_words.xlsx'
wb = openpyxl.load_workbook(xlfile)
wsSheet = wb['pattern6']
max_row = wsSheet.max_row
field = []

for t in range(2,18):
    words = {}
    for i in range(1,max_row+1):
        value = wsSheet.cell(i,t).value
        if value is None:
            continue
        word = wsSheet.cell(i,1).value
        words[word] = value
    field.append(words)
for p in field:
    print(str(p)+", \\")
